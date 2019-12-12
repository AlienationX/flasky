# coding=utf-8
# python3

import openpyxl
from openpyxl.styles.fonts import Font
from openpyxl.styles.named_styles import NamedStyle
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def a():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = 'You should see three logos below'
    ws['A2'] = '中国'
    ws['B1'] = 'You'
    ws['B2'] = '伤害'
    ws.sheet_properties.tabColor = "1072BA"
    # ws._fonts.add(Font(size=8))
    # ws.add_named_style(NamedStyle(font=Font(size=8), builtinId=0))

    for row in ws.rows:
        for cell in row:
            print(cell)
            # Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
            cell.font = Font(size=8)
    wb.save('group.xlsx')


if __name__ == '__main__':
    a()
